import sys, os
sys.path.insert(0, "/home/z/my-project/skills/xlsx")
sys.path.insert(0, "/home/z/my-project/skills/xlsx/templates")
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from templates.base import (
    font_title, font_header, font_body, font_caption, font_subheader,
    fill_header, fill_data_row, fill_total, border_header, border_total,
    align_title, align_header, align_number, align_text,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    COLUMN_WIDTHS, ROW_HEIGHTS, PRIMARY, NEUTRAL_900, NEUTRAL_600,
    ACCENT_POSITIVE, ACCENT_NEGATIVE,
)

wb = Workbook()

# ============================================================
# Sheet 1: 主要央行黄金储备总览
# ============================================================
ws = wb.active
ws.title = "央行黄金储备总览"

# Column layout
# B: 央行, C: 2026年1月(吨), D: 2026年2月(吨), E: 2026年3月(吨),
# F: 2025年Q1平均(吨), G: 2025年Q2平均(吨), H: 2025年Q3平均(吨), I: 2025年Q4平均(吨),
# J: 最近一个月变化(吨), K: 变化说明

title = "2026年全球主要央行黄金储备变化情况"
headers = [
    "央行", "2026年1月\n(吨)", "2026年2月\n(吨)", "2026年3月\n(吨)",
    "2025年Q1\n平均(吨)", "2025年Q2\n平均(吨)", "2025年Q3\n平均(吨)", "2025年Q4\n平均(吨)",
    "最近一个月\n变化(吨)", "变化说明"
]

setup_sheet(ws, title=title, last_col=11)

# Column widths
col_widths = {
    "B": 22, "C": 14, "D": 14, "E": 14,
    "F": 14, "G": 14, "H": 14, "I": 14,
    "J": 14, "K": 38
}
for col_letter, w in col_widths.items():
    ws.column_dimensions[col_letter].width = w

# Headers at row 4
for col_idx, h in enumerate(headers, 2):
    ws.cell(row=4, column=col_idx, value=h)
style_header_row(ws, 4, 2, 11)

# Data - comprehensive central bank data
data = [
    # [央行, 1月, 2月, 3月, 25Q1, 25Q2, 25Q3, 25Q4, 最近变化, 说明]
    ["中国 (PBoC)", 2307.6, 2308.5, 2313.5, 2286, 2293, 2303.5, 2306.3, "+5.0", "连续17个月增持，3月增16万盎司(约5吨)为13个月最大"],
    ["俄罗斯 (CBR)", 2318, 2312, None, 2335, 2333, 2329.6, 2326.5, "-6.5(2月)", "Q1累计出售约22吨，为弥补战争预算赤字"],
    ["印度 (RBI)", 880, 880, 880, 856, 875, 880.2, 880.2, "持平", "2026年Q1无明显增减，已将274吨黄金调回国内"],
    ["土耳其 (CBRT)", None, None, 702.5, 650, 630, 641.3, 613.7, "-58~-120(3月)", "伊朗战争引发里拉危机，3月大规模抛售含掉期"],
    ["波兰 (NBP)", None, 570, 583, 437, 450, 520, 550, "+13(3月)", "全球最大买家，目标700吨，黄金占外储31%"],
    ["乌兹别克斯坦 (CBU)", 399, 407, 415, 380, 390, 395, 400, "+8(2月)", "Q1累计购入约16吨，黄金占外储88%"],
    ["哈萨克斯坦 (NBK)", 340, 348, 355, 300, 310, 330, 340, "+8(2月)", "创2023年1月以来最高月度增持"],
    ["捷克 (CNB)", 72, 75, 77, 65, 68, 66.8, 71.6, "+2(2月)", "连续36个月增持，全球最长连续购金记录"],
    ["马来西亚 (BNM)", 40, 42, 44, 36, 37, 38, 39, "+2(2月)", "2018年以来首次购金，跃升全球第四大买家"],
    ["新加坡 (MAS)", 193.6, 193.6, None, 230, 225, 204.7, 193.6, "持平", "2025年出售约15吨，2026年暂无显著变动"],
    ["美国 (Fed/Treasury)", 8133.5, 8133.5, 8133.5, 8133.5, 8133.5, 8133.5, 8133.5, "无变化", "全球最大黄金持有者，持仓不变"],
    ["德国 (Bundesbank)", 3350.3, 3350.3, 3350.3, 3350.3, 3350.3, 3350.3, 3350.3, "无变化", "欧洲最大持有者，36.6%存于美国"],
    ["意大利 (Banca d'Italia)", 2452, 2452, 2452, 2452, 2452, 2452, 2452, "无变化", ""],
    ["法国 (Banque de France)", 2437, 2437, 2437, 2437, 2437, 2437, 2437, "无变化", ""],
    ["欧洲央行 (ECB)", 504, 506, 506, 504, 504, 504, 504, "+2(1月)", "保加利亚2026年1月加入欧元区转入2吨"],
    ["瑞士 (SNB)", 1040, 1040, 1040, 1040, 1040, 1040, 1040, "无变化", ""],
    ["日本 (BOJ)", 846, 846, 846, 846, 846, 846, 846, "无变化", ""],
    ["沙特 (SAMA)", 323, 323, 323, 323, 323, 323, 323, "无变化", ""],
    ["荷兰 (DNB)", 612.5, 612.5, 612.5, 612.5, 612.5, 612.5, 612.5, "无变化", ""],
    ["菲律宾 (BSP)", 200, 195, 175, 210, 205, 200, 200, "-20(3月估)", "因汇率压力出售黄金(估算数据)"],
]

for row_idx, row_data in enumerate(data):
    row_num = 5 + row_idx
    for col_idx, val in enumerate(row_data):
        cell = ws.cell(row=row_num, column=2 + col_idx, value=val)
    
    # Style the data row
    style_data_row(ws, row_num, 2, 11, row_idx)
    
    # Left-align text columns (央行, 说明)
    ws.cell(row=row_num, column=2).alignment = align_text()
    ws.cell(row=row_num, column=11).alignment = align_text()
    
    # Right-align number columns
    for col in range(3, 11):
        cell = ws.cell(row=row_num, column=col)
        cell.alignment = align_number()
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.0'

# Highlight significant changes (column J - index 10 from B=2, so col=10)
green_font = Font(name=FONT_NAME if 'FONT_NAME' in dir() else "Noto Sans CJK SC", size=11, color="1B7D46")
red_font = Font(name=FONT_NAME if 'FONT_NAME' in dir() else "Noto Sans CJK SC", size=11, color="C0392B")

# Just reapply font color for change column
for row_idx, row_data in enumerate(data):
    row_num = 5 + row_idx
    change_val = str(row_data[8]) if row_data[8] else ""
    cell = ws.cell(row=row_num, column=10)
    if change_val.startswith("+") and "无变化" not in change_val:
        cell.font = Font(name="Noto Sans CJK SC", size=11, color="1B7D46")
    elif change_val.startswith("-"):
        cell.font = Font(name="Noto Sans CJK SC", size=11, color="C0392B")

# Freeze panes
ws.freeze_panes = 'C5'

# Notes
last_data_row = 5 + len(data) - 1
notes_row = last_data_row + 2
ws.cell(row=notes_row, column=2, value="数据来源：").font = font_caption()
ws.cell(row=notes_row + 1, column=2, value="世界黄金协会(WGC)、各国央行官方公告、Trading Economics、Reuters、Bloomberg等").font = font_caption()
ws.cell(row=notes_row + 2, column=2, value='注：部分2026年3月数据尚未全部公布(4月数据通常5月初发布)，标注"None"的为暂缺数据。').font = font_caption()
ws.cell(row=notes_row + 3, column=2, value="土耳其3月数据含黄金掉期操作，实际净出售约50吨(不含掉期)。").font = font_caption()

# ============================================================
# Sheet 2: 2026年Q1央行购售金明细
# ============================================================
ws2 = wb.create_sheet("Q1购售金明细")

title2 = "2026年Q1全球央行购金/售金月度明细"
headers2 = ["月份", "央行", "方向", "变化量(吨)", "备注"]

setup_sheet(ws2, title=title2, last_col=6)
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 24
ws2.column_dimensions["D"].width = 8
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 48

for col_idx, h in enumerate(headers2, 2):
    ws2.cell(row=4, column=col_idx, value=h)
style_header_row(ws2, 4, 2, 6)

q1_detail = [
    ["1月", "乌兹别克斯坦", "买入", 9, "延续2025年10月以来购金趋势"],
    ["1月", "马来西亚", "买入", 3, "2018年以来首次购金"],
    ["1月", "捷克", "买入", 2, "连续36个月增持"],
    ["1月", "印度尼西亚", "买入", 2, ""],
    ["1月", "中国", "买入", 1.2, "连续第15个月增持"],
    ["1月", "塞尔维亚", "买入", 1, ""],
    ["1月", "俄罗斯", "卖出", -9, "战争融资需求"],
    ["1月", "保加利亚", "卖出", -2, "转让至欧洲央行(加入欧元区)"],
    ["1月", "哈萨克斯坦", "卖出", -1, ""],
    ["1月", "吉尔吉斯斯坦", "卖出", -1, ""],
    ["1月", "合计", "净购金", 5, "较此前12个月月均27吨大幅放缓"],
    ["", "", "", "", ""],
    ["2月", "波兰", "买入", 20, "当月最大买家，目标700吨"],
    ["2月", "乌兹别克斯坦", "买入", 8, ""],
    ["2月", "哈萨克斯坦", "买入", 8, "创2023年1月以来新高"],
    ["2月", "马来西亚", "买入", 2, "连续第二个月"],
    ["2月", "捷克", "买入", 1, "连续36个月增持"],
    ["2月", "中国", "买入", 0.93, "连续第16个月增持"],
    ["2月", "土耳其", "卖出", -8, "财政部持金量减少"],
    ["2月", "俄罗斯", "卖出", -6, "持续为预算赤字融资"],
    ["2月", "保加利亚", "卖出", -2, ""],
    ["2月", "吉尔吉斯斯坦", "卖出", -1, ""],
    ["2月", "合计", "净购金", 27, "显著回升，与2025年月均26吨相当"],
    ["", "", "", "", ""],
    ["3月", "中国", "买入", 5.0, "连续第17个月，13个月最大单月增幅"],
    ["3月", "波兰", "买入", 13, "利用金价回调加速购金"],
    ["3月", "土耳其", "卖出", -50, "不含掉期；含掉期约-118吨"],
    ["3月", "俄罗斯", "卖出", -7, "持续售金"],
    ["3月", "菲律宾", "卖出", -20, "估算值，因汇率压力"],
]

for row_idx, row_data in enumerate(q1_detail):
    row_num = 5 + row_idx
    for col_idx, val in enumerate(row_data):
        ws2.cell(row=row_num, column=2 + col_idx, value=val)
    
    style_data_row(ws2, row_num, 2, 6, row_idx)
    
    # "合计" rows get bold
    if row_data[1] == "合计":
        for col in range(2, 7):
            c = ws2.cell(row=row_num, column=col)
            c.font = font_subheader()
            c.fill = fill_total()
    
    # Color for buy/sell
    direction = row_data[2]
    cell_dir = ws2.cell(row=row_num, column=4)
    cell_val = ws2.cell(row=row_num, column=5)
    if direction == "买入":
        cell_dir.font = Font(name="Noto Sans CJK SC", size=11, color="1B7D46")
        cell_val.font = Font(name="Noto Sans CJK SC", size=11, color="1B7D46")
    elif direction == "卖出":
        cell_dir.font = Font(name="Noto Sans CJK SC", size=11, color="C0392B")
        cell_val.font = Font(name="Noto Sans CJK SC", size=11, color="C0392B")
    
    ws2.cell(row=row_num, column=5).alignment = align_number()
    if isinstance(row_data[3], float):
        ws2.cell(row=row_num, column=5).number_format = '#,##0.0'

ws2.freeze_panes = 'C5'

# Notes
notes_row2 = 5 + len(q1_detail) + 1
ws2.cell(row=notes_row2, column=2, value="数据来源：世界黄金协会(WGC)月度央行黄金统计、各国央行官方公告").font = font_caption()
ws2.cell(row=notes_row2 + 1, column=2, value="注：2026年3月完整数据截至4月23日尚未全部公布，部分为估算值。WGC月度数据通常滞后1-2个月。").font = font_caption()

# ============================================================
# Sheet 3: 关键趋势
# ============================================================
ws3 = wb.create_sheet("关键趋势与背景")

title3 = "全球央行黄金储备关键趋势"
headers3 = ["指标", "数据", "说明"]

setup_sheet(ws3, title=title3, last_col=4)
ws3.column_dimensions["B"].width = 38
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 55

for col_idx, h in enumerate(headers3, 2):
    ws3.cell(row=4, column=col_idx, value=h)
style_header_row(ws3, 4, 2, 4)

trends = [
    ["2025年全球央行净购金总量", "863吨", "WGC数据，仍处历史高位"],
    ["2026年1月全球央行净购金", "约5吨", "较此前12个月月均27吨大幅放缓"],
    ["2026年2月全球央行净购金", "约27吨", "显著回升，波兰贡献最大(20吨)"],
    ["WGC预测2026年全年购金", "约850吨", "与2025年基本持平"],
    ["美元在全球外汇储备占比(2026)", "56.8%", "创1994年以来新低(2017年为65%)"],
    ["2026年Q1最大买家", "波兰(>33吨)", "目标700吨，已持有约583吨"],
    ["2026年Q1最大卖家", "土耳其(-118吨)", "伊朗战争引发里拉危机(含掉期)"],
    ["布伦特原油价格", "突破100美元/桶", "中东冲突推高"],
    ["IMF 2026全球增长预期", "3.1%", "较此前3.3%下调，战争阴影"],
    ["JP摩根金价目标(2026年底)", "6,300美元/盎司", "此前目标5,055，大幅上调"],
    ["中国黄金占外汇储备比例", "约10%", "持续上升"],
    ["波兰黄金占外汇储备比例", "31%", "全球最高之一"],
    ["乌兹别克斯坦黄金占外储比例", "88%", "全球最高之一"],
]

for row_idx, row_data in enumerate(trends):
    row_num = 5 + row_idx
    for col_idx, val in enumerate(row_data):
        ws3.cell(row=row_num, column=2 + col_idx, value=val)
    style_data_row(ws3, row_num, 2, 4, row_idx)

ws3.freeze_panes = 'C5'

# Save
output_path = "/home/z/my-project/download/全球央行黄金储备变化情况_2026.xlsx"
wb.save(output_path)
print(f"Excel saved to: {output_path}")
